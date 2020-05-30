#!/usr/bin/env python
# coding: utf-8



import argparse
import pandas as pd
import sys
import numpy as np
from datetime import datetime, timedelta




if len(sys.argv) < 2 or len(sys.argv) < 2:
    print('You need to specify a filename. Like python script.py Input.xlsx')
    sys.exit()

parser = argparse.ArgumentParser(description='Ranker')
parser.add_argument('path',type=str)
args = parser.parse_args()

path = args.path


from openpyxl import load_workbook
workbook = load_workbook(filename=path)




print(path)



dfs = []
for st in workbook.sheetnames:
    print(st)
    sheet = workbook.get_sheet_by_name(st)
    out = []
    for i in sheet:
        temp = []
        for j in i:
            temp.append(j.value)
        out.append(temp)
    name = out.pop(0)[0]
    out.pop(0)
    data = pd.DataFrame(np.array(out)[1:],columns=np.array(out)[0])
    data['Name'] = name
    data.drop(data[data['Broker'] == 'Excluded'].index,inplace=True)
    data.drop(data[data['Broker'] == 'Restricted'].index,inplace=True)
    delete = data[(data == '-').any(axis=1)].index
    data.drop(delete,inplace=True)
    dfs.append(data)
    
    
    

cond1 = pd.concat(dfs).sort_index()
cond1 = cond1[cond1.index == 0].loc[:,['Name','Tgt Price Implied Return (%)']]
cond1['Tgt Price Implied Return (%)'] = cond1['Tgt Price Implied Return (%)'].astype('float64')
cond1.sort_values(by='Tgt Price Implied Return (%)',ascending=False)
cond1.rename(columns={
            'Name':'Ranking based on Mean Target price implied return',
             'Tgt Price Implied Return (%)' : 'Mean Tgt Price Implied Return (%)'  
},inplace=True)
cond1.reset_index(drop=True, inplace=True)



cond2 = pd.concat(dfs).sort_index()
cond2 = cond2[cond2.index == 1].loc[:,['Name','Tgt Price Implied Return (%)']]
cond2['Tgt Price Implied Return (%)'] = cond2['Tgt Price Implied Return (%)'].astype('float64')
cond2.sort_values(by='Tgt Price Implied Return (%)',ascending=False)
cond2.rename(columns={
            'Name':"Ranking based on the most recent rating's target price implied return",
             'Tgt Price Implied Return (%)' : "The Most Recent Rating's Tgt Price Implied Return (%)" 
},inplace=True)
cond2.reset_index(drop=True, inplace=True)



dfs = []
for st in workbook.sheetnames:
    sheet = workbook.get_sheet_by_name(st)
    out = []
    for i in sheet:
        temp = []
        for j in i:
            temp.append(j.value)
        out.append(temp)
    name = out.pop(0)[0]
    out.pop(0)
    data = pd.DataFrame(np.array(out)[1:],columns=np.array(out)[0])
    data['Name'] = name
    data.drop(data[data['Broker'] == 'Excluded'].index,inplace=True)
    data.drop(data[data['Broker'] == 'Restricted'].index,inplace=True)
    delete = data[(data == '-').any(axis=1)].index
    data.drop(delete,inplace=True)
    dfs.append(data)




for i in range(len(dfs)):
    dfs[i]['Rating Date'] = pd.to_datetime(dfs[i]['Rating Date'])
    dfs[i]  = dfs[i][1:].sort_values(by='Rating Date',ascending=False)
    till = (datetime.today() - timedelta(days=21)).strftime('%Y-%m-%d')
    dfs[i] = dfs[i][dfs[i]['Rating Date']>till]
    dfs[i]['Tgt Price Implied Return (%)'] = dfs[i]['Tgt Price Implied Return (%)'].astype('float64')

    dfs[i] = dfs[i].groupby(['Name']).agg({'Tgt Price Implied Return (%)' : ['mean']})
    dfs[i] = dfs[i].reset_index()
    

cond3 = pd.concat(dfs).sort_index()
cond3.columns = ['a', 'b']
cond3.rename(columns={
            'a':"Ranking based on the last 3 weeks' average ratings' target price implied return",
             'b' : "The Last 3 weeks 'avg ratings' target implied return"  
},inplace=True)
cond3.sort_values(by="The Last 3 weeks 'avg ratings' target implied return" ,ascending=False)
cond3.reset_index(drop=True, inplace=True)




dfs = []
for st in workbook.sheetnames:
    sheet = workbook.get_sheet_by_name(st)
    out = []
    for i in sheet:
        temp = []
        for j in i:
            temp.append(j.value)
        out.append(temp)
    name = out.pop(0)[0]
    out.pop(0)
    data = pd.DataFrame(np.array(out)[1:],columns=np.array(out)[0])
    data['Name'] = name
    data.drop(data[data['Broker'] == 'Excluded'].index,inplace=True)
    data.drop(data[data['Broker'] == 'Restricted'].index,inplace=True)
    delete = data[(data == '-').any(axis=1)].index
    data.drop(delete,inplace=True)
    dfs.append(data)




for i in range(len(dfs)):
    dfs[i]['Rating Date'] = pd.to_datetime(dfs[i]['Rating Date'])
    dfs[i]  = dfs[i][1:].sort_values(by='Rating Date',ascending=False)
    till = (datetime.today() - timedelta(days=10)).strftime('%Y-%m-%d')
    dfs[i] = dfs[i][dfs[i]['Rating Date']>till]
    dfs[i]['Tgt Price Implied Return (%)'] = dfs[i]['Tgt Price Implied Return (%)'].astype('float64')
    dfs[i] = dfs[i][dfs[i]['Tgt Price Implied Return (%)'] == dfs[i]['Tgt Price Implied Return (%)'].min()]
cond4 = pd.concat(dfs).sort_index()
cond4 = cond4.loc[:,['Name','Tgt Price Implied Return (%)']]
cond4.sort_values(by='Tgt Price Implied Return (%)',ascending=False)
cond4.rename(columns={
            'Name':"Ranking based on the last 10 days' MIMIMUM ratings' target price implied return",
             'Tgt Price Implied Return (%)' : "The Last 10 days 'min ratings' target implied return"  
},inplace=True)
cond4.reset_index(drop=True, inplace=True)



output = pd.concat([cond1, cond2, cond3, cond4], axis=1, sort=False)
with pd.ExcelWriter('Output.xlsx') as writer:
            output.to_excel(writer,index=False)


